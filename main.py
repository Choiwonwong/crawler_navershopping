from utils import *
from crawling import *
import time, sys
from datetime import timedelta
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from datetime import datetime
from openpyxl import Workbook, load_workbook

driver_path = '/Users/choiwonwong/STUDY/NS_CRAWL/chromedriver-mac-arm64/chromedriver'
output_name = datetime.today().strftime("%Y-%m-%d")

if __name__ == "__main__":
    if sys.argv[-1] == sys.argv[0]:
        print("쿼리 시트가 입력되지 않았습니다.")
        print("사용 방법은 main.py [쿼리 시트 명]입니다. 다시 시도해주세요")
        exit(1)
    elif len(sys.argv) >= 3:
        print("쿼리 시트 지정이 정확하지 않습니다.")
        print("입력된 매개 변수의 수가", len(sys.argv) - 1, "개입니다.")
        print("사용 방법은 main.py [쿼리 시트 명]입니다. 다시 시도해주세요")
        exit(1)
    else:
        fileName = sys.argv[1]
        queryPath = "queries/" + fileName + '.xlsx'
        try:
            print(fileName, "쿼리 시트에서 쿼리를 읽습니다.")
            querieXlsx = load_workbook(queryPath, data_only=True)
            queriesSheet = querieXlsx['시트1']
            queriesSheet.delete_rows(0)
            queries = []
            for query, pageCount in queriesSheet.rows:
                queries.append([query.value, int(pageCount.value)])
            print(len(queries), "개의 쿼리가 정상적으로 확인되었습니다.")
        except:
            print(fileName, "쿼리 시트에서 쿼리를 읽을 수 없습니다.")
            print("파일 이름 또는 시트 구조를 확인해주세요")

    start_time = time.time()
    driver = chrome_driver(driver_path)
    shopping = setupBeforeCrawling(chromeDriver=driver)

    # 입력 버튼 선택
    search = shopping.find_element(By.XPATH, "/html/body/div[3]/div/div[1]/div/div/div/div[2]/div/div[2]/div/div[2]/form/div[1]/div/input") # XPATH
    search.click()

    crawledItems = []

    for query, pageCount in queries:
        # 쿼리 검색
        search.send_keys(query)
        search.send_keys(Keys.ENTER)
        shopping.implicitly_wait(3)
        adCount = 0
        count = 0

        subDriver = chrome_driver(driver_path, headless=True)
        print("===========[{}] 데이터를 수집합니다.".format(query))
        for pageIdx in range(1, pageCount + 1):
            time.sleep(2)
            # 무한 스크롤 처리
            processInfiniteScroll(shopping)

            # 전체 상품 리스트 - div > basicList_basis__uNBZx
            items = shopping.find_element(By.CLASS_NAME, "basicList_list_basis__uNBZx")

            # 광고 제품 HTML 요소 - div > adProduct_inner__W_nuz
            adElements = items.find_elements(By.CLASS_NAME, "adProduct_inner__W_nuz")
            adItems = crawlAdItems(adElements=adElements, subDriver=subDriver)

            crawledItems += adItems
            adCount += len(adItems)
            
            # 미광고 제품 - div > product_inner__gr8QR
            # elements = items.find_elements(By.CLASS_NAME, "product_inner__gr8QR")
            # items = crawlItems(elements=elements, subDriver=subDriver)

            # crawledItems += items
            try: 
                shopping.find_element(By.CLASS_NAME, 'pagination_next__pZuC6').click()
                print("===========[{}] - [{}]".format(query, pageIdx))
            except:
                print("===========[{}]가 마지막 페이지입니다.".format(pageIdx))
                break

        print("===========[{}] 데이터 [{}] 건이 수집되었습니다.([{}]p)".format(query, adCount, pageCount))
        # print("========= 수집된 데이터 개수: ", len(adItems) + len(items),"===========")

        search = shopping.find_element(By.XPATH, "/html/body/div/div/div[1]/div[2]/div/div[2]/div/div[2]/form/div[1]/div/input") # XPATH
        search.send_keys(Keys.COMMAND, 'a') # 맥북이라 COMMAND인 듯
        search.send_keys(Keys.DELETE)
        search.click() 

    wb = Workbook()
    ws = wb.active
    
    headers = list(crawledItems[0].keys())

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    for row_idx, data_dict in enumerate(crawledItems, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=data_dict.get(header))
    
    wb.save("results/" + output_name + ".xlsx")
    print("[" + output_name + "] 크롤링 결과물이 생성되었습니다.")
    end_time = time.time()
    sec = end_time - start_time
    print("전체 소요 시간:", timedelta(seconds=sec))